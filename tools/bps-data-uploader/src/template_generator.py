#!/usr/bin/env python3
"""
BPS Data Template Generator
Creates the master Excel template for BPS staff to fill in statistical data.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any

class BPSTemplateGenerator:
    """Generates Excel template matching BPS workflow patterns."""
    
    def __init__(self, output_path: str):
        self.output_path = output_path
        self.workbook = openpyxl.Workbook()
        self.years = ['2020', '2021', '2022', '2023', '2024', '2025', '2026']
        
        # Styles
        self.header_font = Font(bold=True, size=11, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def create_instruction_sheet(self):
        """Create instructions sheet."""
        ws = self.workbook.active
        ws.title = "Petunjuk"
        
        instructions = [
            ["BPS DATA UPLOADER - PETUNJUK PENGGUNAAN"],
            [""],
            ["1. CARA MENGGUNAKAN"],
            ["   - Buka sheet kategori yang ingin diupdate"],
            ["   - Isi data pada kolom tahun yang sesuai"],
            ["   - Kosongkan sel yang tidak ada datanya"],
            ["   - Simpan file (Ctrl+S)"],
            ["   - Jalankan BPS Data Uploader"],
            [""],
            ["2. DAFTAR KATEGORI"],
            ["   Sheet 1: Petunjuk (panduan ini)"],
            ["   Sheet 2: Pertumbuhan_Ekonomi - Data pertumbuhan ekonomi"],
            ["   Sheet 3: Inflasi - Data inflasi bulanan"],
            ["   Sheet 4: Tenaga_Kerja - Data ketenagakerjaan"],
            ["   Sheet 5: Kemiskinan - Data kemiskinan"],
            ["   Sheet 6: Penduduk - Data penduduk"],
            ["   Sheet 7: Pendidikan - Data pendidikan"],
            ["   Sheet 8: IPM - Indeks Pembangunan Manusia"],
            ["   Sheet 9: IPG - Indeks Pembangunan Gender"],
            ["   Sheet 10: IDG - Indeks Ketimpangan Gender"],
            ["   Sheet 11: SDGs - Sustainable Development Goals"],
            [""],
            ["3. KETERANGAN WARNA"],
            ["   Biru tua: Header/keterangan"],
            ["   Biru muda: Sub-kategori"],
            ["   Putih: Isian data"],
            [""],
            ["4. VALIDASI DATA"],
            ["   - Persentase tidak boleh > 100%"],
            ["   - Jumlah penduduk dalam ribuan"],
            ["   - Gunakan titik (.) untuk desimal, bukan koma"],
            [""],
            ["5. BANTUAN"],
            ["   Hubungi admin IT BPS jika ada kendala"],
        ]
        
        for row_idx, row_data in enumerate(instructions, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14, color="4472C4")
                elif "CARA" in str(value) or "DAFTAR" in str(value) or "KETERANGAN" in str(value) or "VALIDASI" in str(value) or "BANTUAN" in str(value):
                    cell.font = Font(bold=True, size=12)
        
        ws.column_dimensions['A'].width = 80
        
    def create_pertumbuhan_ekonomi_sheet(self):
        """Create Pertumbuhan Ekonomi sheet."""
        ws = self.workbook.create_sheet("Pertumbuhan_Ekonomi")
        
        # Headers
        headers = ["Indikator"] + self.years
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data rows
        indicators = [
            "Pertumbuhan Ekonomi (%)",
            "Kontribusi PDRB (%)",
            "Sektor Perdagangan (%)",
            "PDRB Per Kapita (Juta Rupiah)",
            "Rank di Jawa Tengah",
            "TPT Nasional (%)"
        ]
        
        for row_idx, indicator in enumerate(indicators, 2):
            cell = ws.cell(row=row_idx, column=1, value=indicator)
            cell.border = self.border
            for col in range(2, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
            
    def create_inflasi_sheet(self):
        """Create Inflasi sheet - Monthly data format."""
        ws = self.workbook.create_sheet("Inflasi")
        
        # Headers
        headers = ["Komponen"] + self.years
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Components
        components = [
            "Inflasi Umum (%)",
            "Inflasi Inti (%)",
            "IHK (Indeks)",
            "Makanan, Minuman & Tembakau",
            "Pakaian & Alas Kaki",
            "Perumahan & Fasilitas",
            "Perlengkapan & Perawatan Rumah",
            "Kesehatan",
            "Transportasi",
            "Informasi, Komunikasi & Keuangan",
            "Rekreasi, Olahraga & Budaya",
            "Pendidikan",
            "Penyediaan Makanan & Minuman",
            "Perawatan Pribadi & Jasa Lainnya"
        ]
        
        for row_idx, component in enumerate(components, 2):
            cell = ws.cell(row=row_idx, column=1, value=component)
            cell.border = self.border
            for col in range(2, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 40
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
            
    def create_tenaga_kerja_sheet(self):
        """Create Tenaga Kerja sheet."""
        ws = self.workbook.create_sheet("Tenaga_Kerja")
        
        headers = ["Indikator"] + self.years
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        indicators = [
            "Angkatan Kerja (Ribu)",
            "Bekerja (Ribu)",
            "Pengangguran (Ribu)",
            "Bukan Angkatan Kerja (Ribu)",
            "Tingkat Pengangguran Terbuka - TPT (%)",
            "Tingkat Kesempatan Kerja - TKK (%)",
            "Tingkat Partisipasi Angkatan Kerja - TPAK (%)",
            "TPT Laki-laki (%)",
            "TPT Perempuan (%)",
            "TKK Laki-laki (%)",
            "TKK Perempuan (%)",
            "TPAK Laki-laki (%)",
            "TPAK Perempuan (%)",
            "Sektor Pertanian (%)",
            "Sektor Manufaktur (%)",
            "Sektor Jasa (%)"
        ]
        
        for row_idx, indicator in enumerate(indicators, 2):
            cell = ws.cell(row=row_idx, column=1, value=indicator)
            cell.border = self.border
            for col in range(2, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 40
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
            
    def create_kemiskinan_sheet(self):
        """Create Kemiskinan sheet."""
        ws = self.workbook.create_sheet("Kemiskinan")
        
        headers = ["Indikator"] + self.years
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        indicators = [
            "Jumlah Penduduk Miskin (Ribu)",
            "Persentase Penduduk Miskin (%)",
            "Garis Kemiskinan (Rupiah/Bulan)",
            "Indeks Kedalaman Kemiskinan (P1)",
            "Indeks Keparahan Kemiskinan (P2)"
        ]
        
        for row_idx, indicator in enumerate(indicators, 2):
            cell = ws.cell(row=row_idx, column=1, value=indicator)
            cell.border = self.border
            for col in range(2, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 40
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
            
    def create_penduduk_sheet(self):
        """Create Penduduk sheet."""
        ws = self.workbook.create_sheet("Penduduk")
        
        headers = ["Indikator"] + self.years
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        indicators = [
            "Jumlah Penduduk (Ribu)",
            "Penduduk Laki-laki (Ribu)",
            "Penduduk Perempuan (Ribu)",
            "Luas Wilayah (km2)",
            "Kepadatan Penduduk (jiwa/km2)",
            "Jumlah Kecamatan",
            "Jumlah Kelurahan/Desa",
            "Laju Pertumbuhan Penduduk (%)",
            "Persentase Usia Muda (0-14 tahun) (%)",
            "Persentase Usia Produktif (15-64 tahun) (%)",
            "Persentase Usia Tua (65+ tahun) (%)"
        ]
        
        for row_idx, indicator in enumerate(indicators, 2):
            cell = ws.cell(row=row_idx, column=1, value=indicator)
            cell.border = self.border
            for col in range(2, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 45
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
            
    def create_pendidikan_sheet(self):
        """Create Pendidikan sheet."""
        ws = self.workbook.create_sheet("Pendidikan")
        
        headers = ["Indikator"] + self.years
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        indicators = [
            "Angka Melek Huruf - AMH (%)",
            "Rata-rata Lama Sekolah - RLS (Tahun)",
            "Harapan Lama Sekolah - HLS (Tahun)",
            "Rasio Guru-Murid",
            "Tingkat Kelulusan (%)",
            "Akses Pendidikan Tinggi (%)",
            "APM SD/MI (%)",
            "APM SMP/MTs (%)",
            "APM SMA/SMK/MA (%)",
            "APK SD/MI (%)",
            "APK SMP/MTs (%)",
            "APK SMA/SMK/MA (%)"
        ]
        
        for row_idx, indicator in enumerate(indicators, 2):
            cell = ws.cell(row=row_idx, column=1, value=indicator)
            cell.border = self.border
            for col in range(2, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 40
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
            
    def create_ipm_sheet(self):
        """Create IPM sheet."""
        ws = self.workbook.create_sheet("IPM")
        
        headers = ["Indikator"] + self.years
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        indicators = [
            "Usia Harapan Hidup - UHH (Tahun)",
            "Rata-rata Lama Sekolah - RLS (Tahun)",
            "Harapan Lama Sekolah - HLS (Tahun)",
            "Pengeluaran per Kapita (Ribu Rupiah)",
            "Indeks Pembangunan Manusia - IPM",
            "IPM Nasional",
            "IPM Jawa Tengah",
            "IPM Kota Semarang"
        ]
        
        for row_idx, indicator in enumerate(indicators, 2):
            cell = ws.cell(row=row_idx, column=1, value=indicator)
            cell.border = self.border
            for col in range(2, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 40
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
            
    def create_ipg_sheet(self):
        """Create IPG sheet."""
        ws = self.workbook.create_sheet("IPG")
        
        headers = ["Indikator"] + self.years
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        indicators = [
            "IPG Laki-laki",
            "IPG Perempuan",
            "IPG Total",
            "RLS Laki-laki (Tahun)",
            "RLS Perempuan (Tahun)",
            "Pengeluaran Laki-laki (Ribu Rupiah)",
            "Pengeluaran Perempuan (Ribu Rupiah)"
        ]
        
        for row_idx, indicator in enumerate(indicators, 2):
            cell = ws.cell(row=row_idx, column=1, value=indicator)
            cell.border = self.border
            for col in range(2, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 40
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
            
    def create_idg_sheet(self):
        """Create IDG sheet."""
        ws = self.workbook.create_sheet("IDG")
        
        headers = ["Indikator"] + self.years
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        indicators = [
            "Sumbangan Pendapatan Perempuan (%)",
            "Perempuan sebagai Tenaga Profesional (%)",
            "Keterlibatan Perempuan di Parlemen (%)",
            "Indeks Pemberdayaan Gender - IDG",
            "Indeks Ketimpangan Gender - IKG"
        ]
        
        for row_idx, indicator in enumerate(indicators, 2):
            cell = ws.cell(row=row_idx, column=1, value=indicator)
            cell.border = self.border
            for col in range(2, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 45
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
            
    def create_sdgs_sheet(self):
        """Create SDGs sheet - 35 cities, 6 indicators."""
        ws = self.workbook.create_sheet("SDGs")
        
        # Headers
        headers = ["Kota/Kabupaten", "Indikator"] + self.years
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # 35 cities in Jawa Tengah
        cities = [
            "Kab. Cilacap", "Kab. Banyumas", "Kab. Purbalingga", "Kab. Banjarnegara",
            "Kab. Kebumen", "Kab. Purworejo", "Kab. Wonosobo", "Kab. Magelang",
            "Kab. Boyolali", "Kab. Klaten", "Kab. Sukoharjo", "Kab. Wonogiri",
            "Kab. Karanganyar", "Kab. Sragen", "Kab. Grobogan", "Kab. Blora",
            "Kab. Rembang", "Kab. Pati", "Kab. Kudus", "Kab. Jepara",
            "Kab. Demak", "Kab. Semarang", "Kab. Temanggung", "Kab. Kendal",
            "Kab. Batang", "Kab. Pekalongan", "Kab. Pemalang", "Kab. Tegal",
            "Kab. Brebes", "Kota Magelang", "Kota Surakarta", "Kota Salatiga",
            "Kota Semarang", "Kota Pekalongan", "Kota Tegal"
        ]
        
        indicators = [
            "Samitasilayak (%)",
            "TIK Remaja (%)",
            "TIK Dewasa (%)",
            "Akta Lahir (%)",
            "APM (%)",
            "APK (%)"
        ]
        
        row = 2
        for city in cities:
            # City header (merged)
            ws.cell(row=row, column=1, value=city)
            ws.cell(row=row, column=1).font = Font(bold=True, size=10)
            ws.cell(row=row, column=1).fill = self.subheader_fill
            
            for indicator in indicators:
                ws.cell(row=row, column=2, value=indicator)
                ws.cell(row=row, column=2).border = self.border
                for col in range(3, len(headers) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = self.border
                    cell.alignment = Alignment(horizontal='center')
                row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        for col in range(3, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 10
    
    def generate(self):
        """Generate complete template."""
        print("Generating BPS Master Data Template...")
        
        self.create_instruction_sheet()
        self.create_pertumbuhan_ekonomi_sheet()
        self.create_inflasi_sheet()
        self.create_tenaga_kerja_sheet()
        self.create_kemiskinan_sheet()
        self.create_penduduk_sheet()
        self.create_pendidikan_sheet()
        self.create_ipm_sheet()
        self.create_ipg_sheet()
        self.create_idg_sheet()
        self.create_sdgs_sheet()
        
        self.workbook.save(self.output_path)
        print(f"✅ Template created: {self.output_path}")
        print(f"   Sheets: {len(self.workbook.sheetnames)}")
        print(f"   Categories: 10")
        print(f"   Years: {', '.join(self.years)}")
        print(f"   Cities (SDGs): 35")

if __name__ == "__main__":
    generator = BPSTemplateGenerator("BPS_Master_Data_Template.xlsx")
    generator.generate()
