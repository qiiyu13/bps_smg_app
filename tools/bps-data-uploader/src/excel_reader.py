#!/usr/bin/env python3
"""
BPS Excel Reader Module
Reads BPS-format Excel files and converts to Python data structures.
"""

import pandas as pd
import openpyxl
from typing import Dict, List, Any, Optional
from dataclasses import dataclass
from datetime import datetime

@dataclass
class ValidationResult:
    """Result of data validation."""
    is_valid: bool
    errors: List[str]
    warnings: List[str]

class BPSExcelReader:
    """Reads BPS Master Data Excel files."""
    
    # Category mapping
    CATEGORIES = {
        'Pertumbuhan_Ekonomi': 'pertumbuhan_ekonomi',
        'Inflasi': 'inflasi',
        'Tenaga_Kerja': 'tenaga_kerja',
        'Kemiskinan': 'kemiskinan',
        'Penduduk': 'penduduk',
        'Pendidikan': 'pendidikan',
        'IPM': 'ipm',
        'IPG': 'ipg',
        'IDG': 'idg',
        'SDGs': 'sdgs'
    }
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.data = {}
        
    def load(self) -> bool:
        """Load Excel file."""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            return True
        except Exception as e:
            print(f"Error loading Excel file: {e}")
            return False
    
    def read_all_categories(self) -> Dict[str, Any]:
        """Read all category sheets."""
        if not self.workbook:
            if not self.load():
                return {}
        
        for sheet_name, category_key in self.CATEGORIES.items():
            if sheet_name in self.workbook.sheetnames:
                print(f"Reading {sheet_name}...")
                self.data[category_key] = self._read_category_sheet(sheet_name)
            else:
                print(f"Sheet {sheet_name} not found, skipping...")
        
        return self.data
    
    def _read_category_sheet(self, sheet_name: str) -> Dict[str, Any]:
        """Read a single category sheet."""
        ws = self.workbook[sheet_name]
        
        # Get headers (first row)
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value))
        
        # Years are columns after the first (indicator name column)
        years = headers[1:] if headers else []
        
        # Read data rows
        indicators_data = {}
        
        if sheet_name == 'SDGs':
            # Special handling for SDGs (city, indicator, year columns)
            return self._read_sdgs_sheet(ws)
        
        # Standard format: indicator row, years as columns
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            
            indicator_name = str(row[0]).strip()
            year_values = {}
            
            for idx, year in enumerate(years, start=1):
                if idx < len(row) and row[idx] is not None:
                    try:
                        # Convert to float, handling both comma and dot decimals
                        value = str(row[idx]).replace(',', '.')
                        year_values[year] = float(value)
                    except (ValueError, TypeError):
                        year_values[year] = None
            
            if year_values:
                indicators_data[indicator_name] = year_values
        
        return {
            'sheet_name': sheet_name,
            'years': years,
            'data': indicators_data,
            'last_updated': datetime.now().isoformat()
        }
    
    def _read_sdgs_sheet(self, ws) -> Dict[str, Any]:
        """Read SDGs sheet with 35 cities."""
        cities_data = {}
        current_city = None
        
        years = []
        for cell in ws[1][2:]:  # Skip first 2 columns (Kota, Indikator)
            if cell.value:
                years.append(str(cell.value))
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3:
                continue
            
            # Column 0: City name (if not empty, it's a new city)
            # Column 1: Indicator name
            # Columns 2+: Year values
            
            if row[0] and str(row[0]).strip():
                # New city
                current_city = str(row[0]).strip()
                cities_data[current_city] = {}
            
            if current_city and row[1]:
                indicator = str(row[1]).strip()
                year_values = {}
                
                for idx, year in enumerate(years, start=2):
                    if idx < len(row) and row[idx] is not None:
                        try:
                            value = str(row[idx]).replace(',', '.')
                            year_values[year] = float(value)
                        except (ValueError, TypeError):
                            year_values[year] = None
                
                if indicator not in cities_data[current_city]:
                    cities_data[current_city][indicator] = {}
                cities_data[current_city][indicator] = year_values
        
        return {
            'sheet_name': 'SDGs',
            'years': years,
            'cities': cities_data,
            'city_count': len(cities_data),
            'last_updated': datetime.now().isoformat()
        }
    
    def validate_data(self, category: str) -> ValidationResult:
        """Validate data for a category."""
        errors = []
        warnings = []
        
        if category not in self.data:
            return ValidationResult(False, [f"Category {category} not found"], [])
        
        data = self.data[category]
        
        if category == 'sdgs':
            return self._validate_sdgs(data)
        
        # Validate standard categories
        for indicator, year_values in data.get('data', {}).items():
            # Check for values > 100% (likely error for percentages)
            for year, value in year_values.items():
                if value is not None:
                    if value > 150:
                        warnings.append(f"{indicator} {year}: Value {value} seems very high")
                    if value > 100 and 'persen' in indicator.lower():
                        warnings.append(f"{indicator} {year}: Percentage > 100% ({value})")
                    if value < 0:
                        errors.append(f"{indicator} {year}: Negative value ({value})")
        
        is_valid = len(errors) == 0
        return ValidationResult(is_valid, errors, warnings)
    
    def _validate_sdgs(self, data: Dict) -> ValidationResult:
        """Validate SDGs data."""
        errors = []
        warnings = []
        
        for city, indicators in data.get('cities', {}).items():
            for indicator, year_values in indicators.items():
                for year, value in year_values.items():
                    if value is not None:
                        if value > 200:
                            warnings.append(f"{city} - {indicator} {year}: Value {value} seems very high")
                        if value > 100 and ('APM' in indicator or 'Akta Lahir' in indicator):
                            warnings.append(f"{city} - {indicator} {year}: Percentage > 100% ({value})")
                        if value < 0:
                            errors.append(f"{city} - {indicator} {year}: Negative value ({value})")
        
        is_valid = len(errors) == 0
        return ValidationResult(is_valid, errors, warnings)
    
    def get_category_summary(self, category: str) -> str:
        """Get summary of a category."""
        if category not in self.data:
            return f"Category {category}: No data"
        
        data = self.data[category]
        years = data.get('years', [])
        
        if category == 'sdgs':
            cities = data.get('cities', {})
            return f"SDGs: {len(cities)} cities, {len(years)} years"
        
        indicators = data.get('data', {})
        return f"{category}: {len(indicators)} indicators, {len(years)} years"

if __name__ == "__main__":
    # Test reader
    reader = BPSExcelReader("BPS_Master_Data_Template.xlsx")
    data = reader.read_all_categories()
    
    print("\n" + "="*60)
    print("DATA SUMMARY")
    print("="*60)
    
    for category in reader.CATEGORIES.values():
        if category in data:
            print(reader.get_category_summary(category))
            result = reader.validate_data(category)
            if result.warnings:
                print(f"  Warnings: {len(result.warnings)}")
